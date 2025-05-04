import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import uuid
from models import Staff, User, Vehicle
import tkinter as tk
from tkinter import filedialog, messagebox

class ExportManager:
    """送迎スケジュールをエクスポートするためのクラス"""
    
    def __init__(self, app):
        self.app = app
        self.export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "exports")
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_to_excel(self, routes):
        """送迎スケジュールをExcelにエクスポート"""
        if not routes:
            return None
            
        # 新しいワークブックを作成
        wb = Workbook()
        
        # 曜日ごとにシートを作成
        for day in self.app.settings["workdays"]:
            # 朝のルート
            morning_routes = [r for r in routes if r.date == day and r.is_morning]
            if morning_routes:
                ws = wb.create_sheet(title=f"{day}_朝")
                self._fill_excel_sheet(ws, morning_routes, day, True)
            
            # 夕方のルート
            evening_routes = [r for r in routes if r.date == day and not r.is_morning]
            if evening_routes:
                ws = wb.create_sheet(title=f"{day}_夕")
                self._fill_excel_sheet(ws, evening_routes, day, False)
        
        # デフォルトシートを削除
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # ファイル名
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(self.export_dir, f"送迎スケジュール_{timestamp}.xlsx")
        
        # ファイルを保存
        wb.save(filepath)
        return filepath
    
    def _fill_excel_sheet(self, ws, routes, day, is_morning):
        """Excelシートにデータを入力"""
        time_of_day = "朝" if is_morning else "夕方"
        
        # ヘッダー
        ws.merge_cells('A1:F1')
        header_cell = ws['A1']
        header_cell.value = f"{day}曜日 {time_of_day}の送迎スケジュール"
        header_cell.font = Font(size=14, bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        
        # ヘッダーの背景色
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
        
        # 各車両のルート
        row = 3
        for route in routes:
            # 車両情報ヘッダー
            ws.merge_cells(f'A{row}:F{row}')
            vehicle_cell = ws[f'A{row}']
            vehicle_cell.value = f"車両: {route.vehicle.name} （運転: {route.driver.name if route.driver else '未定'}, 同乗: {route.assistant.name if route.assistant else 'なし'}）"
            vehicle_cell.font = Font(bold=True)
            vehicle_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # カラムヘッダー
            row += 1
            headers = ["順番", "時間", "種別", "利用者", "住所", "備考"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # ルート詳細
            for i, stop in enumerate(route.stops):
                row += 1
                
                # データの入力
                ws.cell(row=row, column=1).value = i + 1  # 順番
                ws.cell(row=row, column=2).value = stop.time  # 時間
                ws.cell(row=row, column=3).value = "迎え" if stop.is_pickup else "送り"  # 種別
                ws.cell(row=row, column=4).value = stop.user.name if stop.user else "施設"  # 利用者
                ws.cell(row=row, column=5).value = stop.user.address if stop.user else self.app.settings.get("facility_address", "")  # 住所
                
                # 備考（利用者の特記事項）
                if stop.user and hasattr(stop.user, 'notes') and stop.user.notes:
                    ws.cell(row=row, column=6).value = stop.user.notes
            
            # 空行を追加
            row += 2
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 8  # 順番
        ws.column_dimensions['B'].width = 12  # 時間
        ws.column_dimensions['C'].width = 8  # 種別
        ws.column_dimensions['D'].width = 20  # 利用者
        ws.column_dimensions['E'].width = 40  # 住所
        ws.column_dimensions['F'].width = 30  # 備考
    
    def export_to_text(self, routes):
        """送迎スケジュールをテキストにエクスポート（ChatGPTチェック用）"""
        if not routes:
            return None
            
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(self.export_dir, f"送迎スケジュール_ChatGPT用_{timestamp}.txt")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            # ヘッダー情報
            f.write("# 送迎スケジュールチェックプロンプト\n\n")
            f.write("以下の送迎スケジュールをチェックして、問題点や改善点があれば指摘してください。\n\n")
            f.write("## 施設情報\n")
            f.write(f"施設住所: {self.app.settings.get('facility_address', '未設定')}\n\n")
            
            # 各曜日のスケジュールを出力
            f.write("## 送迎スケジュール詳細\n\n")
            
            for day in self.app.settings["workdays"]:
                day_routes = [r for r in routes if r.date == day]
                if day_routes:
                    f.write(f"### {day}曜日\n\n")
                    
                    # 朝の送迎
                    morning_routes = [r for r in day_routes if r.is_morning]
                    if morning_routes:
                        f.write("#### 朝の送迎\n\n")
                        for route in morning_routes:
                            self._write_route_text(f, route)
                    
                    # 夕方の送迎
                    evening_routes = [r for r in day_routes if not r.is_morning]
                    if evening_routes:
                        f.write("#### 夕方の送迎\n\n")
                        for route in evening_routes:
                            self._write_route_text(f, route)
            
            # 分析の観点を追加
            f.write("\n## チェックの観点\n")
            f.write("1. 各車両の利用者数は適切か（過剰に多くないか）\n")
            f.write("2. 各ルートの時間配分は適切か\n")
            f.write("3. 地理的に効率的なルートになっているか\n")
            f.write("4. 特記事項に対応できているか\n")
            f.write("5. 運転手と同乗スタッフの配置は適切か\n")
            f.write("6. 全体的な効率性と安全性のバランスは取れているか\n")
        
        return filepath
    
    def _write_route_text(self, file, route):
        """テキストファイルにルート情報を書き込む"""
        file.write(f"* 車両: {route.vehicle.name} (乗車可能人数: {route.vehicle.capacity}人)\n")
        file.write(f"* 運転手: {route.driver.name if route.driver else '未割り当て'}\n")
        file.write(f"* 同乗スタッフ: {route.assistant.name if route.assistant else 'なし'}\n")
        file.write("\n| 順番 | 時間 | 種別 | 利用者 | 住所 | 備考 |\n")
        file.write("|------|------|------|--------|------|------|\n")
        
        for i, stop in enumerate(route.stops):
            pickup_str = "迎え" if stop.is_pickup else "送り"
            user_name = stop.user.name if stop.user else "施設"
            address = stop.user.address if stop.user else self.app.settings.get("facility_address", "")
            notes = stop.user.notes if stop.user and hasattr(stop.user, 'notes') else ""
            
            file.write(f"| {i+1} | {stop.time} | {pickup_str} | {user_name} | {address} | {notes} |\n")
        
        file.write("\n")

    # 職員、利用者、車両データのエクスポート機能
    def export_data_to_excel(self, data_type):
        """指定したデータ型（職員、利用者、車両）をExcelにエクスポートする"""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if data_type == "staff":
            filename = f"職員データ_{timestamp}.xlsx"
            data_list = self.app.staff_list
            headers = ["ID", "名前", "運転可能", "勤務曜日"]
            
            wb = Workbook()
            ws = wb.active
            ws.title = "職員データ"
            
            # ヘッダーの設定
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # データの書き込み
            for row, staff in enumerate(data_list, 2):
                ws.cell(row=row, column=1).value = staff.id
                ws.cell(row=row, column=2).value = staff.name
                ws.cell(row=row, column=3).value = "はい" if staff.can_drive else "いいえ"
                ws.cell(row=row, column=4).value = ", ".join(staff.workdays)
            
            # 列幅の調整
            ws.column_dimensions['A'].width = 36  # ID
            ws.column_dimensions['B'].width = 20  # 名前
            ws.column_dimensions['C'].width = 10  # 運転可能
            ws.column_dimensions['D'].width = 30  # 勤務曜日
            
        elif data_type == "user":
            filename = f"利用者データ_{timestamp}.xlsx"
            data_list = self.app.user_list
            headers = ["ID", "名前", "住所", "朝-迎え時間", "朝-到着時間", "夕-迎え時間", "夕-到着時間", "制約条件", "通所曜日"]
            
            wb = Workbook()
            ws = wb.active
            ws.title = "利用者データ"
            
            # ヘッダーの設定
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # データの書き込み
            for row, user in enumerate(data_list, 2):
                ws.cell(row=row, column=1).value = user.id
                ws.cell(row=row, column=2).value = user.name
                ws.cell(row=row, column=3).value = user.address
                ws.cell(row=row, column=4).value = user.pickup_time_morning
                ws.cell(row=row, column=5).value = user.dropoff_time_morning
                ws.cell(row=row, column=6).value = user.pickup_time_evening
                ws.cell(row=row, column=7).value = user.dropoff_time_evening
                ws.cell(row=row, column=8).value = user.constraints
                ws.cell(row=row, column=9).value = ", ".join(user.attendance_days)
            
            # 列幅の調整
            ws.column_dimensions['A'].width = 36  # ID
            ws.column_dimensions['B'].width = 20  # 名前
            ws.column_dimensions['C'].width = 40  # 住所
            ws.column_dimensions['D'].width = 15  # 朝-迎え時間
            ws.column_dimensions['E'].width = 15  # 朝-到着時間
            ws.column_dimensions['F'].width = 15  # 夕-迎え時間
            ws.column_dimensions['G'].width = 15  # 夕-到着時間
            ws.column_dimensions['H'].width = 30  # 制約条件
            ws.column_dimensions['I'].width = 30  # 通所曜日
            
        elif data_type == "vehicle":
            filename = f"車両データ_{timestamp}.xlsx"
            data_list = self.app.vehicle_list
            headers = ["ID", "車両名", "乗車可能人数"]
            
            wb = Workbook()
            ws = wb.active
            ws.title = "車両データ"
            
            # ヘッダーの設定
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # データの書き込み
            for row, vehicle in enumerate(data_list, 2):
                ws.cell(row=row, column=1).value = vehicle.id
                ws.cell(row=row, column=2).value = vehicle.name
                ws.cell(row=row, column=3).value = vehicle.capacity
            
            # 列幅の調整
            ws.column_dimensions['A'].width = 36  # ID
            ws.column_dimensions['B'].width = 20  # 車両名
            ws.column_dimensions['C'].width = 15  # 乗車可能人数
        else:
            return None
        
        # ファイルを保存
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)
        return filepath

    def import_data_from_excel(self, data_type):
        """指定したデータ型（職員、利用者、車両）をExcelからインポートする"""
        # ファイル選択ダイアログ
        filetypes = [("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        filepath = filedialog.askopenfilename(
            title=f"{data_type}データのインポート",
            filetypes=filetypes,
            initialdir=self.export_dir
        )
        
        if not filepath:
            return False
        
        try:
            # Excelファイルの読み込み
            wb = load_workbook(filepath)
            ws = wb.active
            
            # 既存データのバックアップ
            if data_type == "staff":
                original_data = self.app.staff_list.copy()
            elif data_type == "user":
                original_data = self.app.user_list.copy()
            elif data_type == "vehicle":
                original_data = self.app.vehicle_list.copy()
            
            # ヘッダー行を飛ばす
            rows = list(ws.rows)
            data_rows = rows[1:]  # 1行目（ヘッダー）をスキップ
            
            if data_type == "staff":
                # 新しいデータリストを作成
                new_data = []
                
                for row in data_rows:
                    # IDは既存のものを保持するか、新規に生成する
                    id_value = row[0].value if row[0].value else str(uuid.uuid4())
                    name = row[1].value
                    can_drive = True if row[2].value == "はい" else False
                    workdays = [day.strip() for day in (row[3].value or "").split(",")] if row[3].value else []
                    
                    # 必須項目のチェック
                    if not name:
                        continue
                    
                    # 新しい職員オブジェクトを作成
                    staff = Staff(
                        id=id_value,
                        name=name,
                        can_drive=can_drive,
                        workdays=workdays
                    )
                    new_data.append(staff)
                
                # データを更新
                self.app.staff_list = new_data
                
            elif data_type == "user":
                # 新しいデータリストを作成
                new_data = []
                
                for row in data_rows:
                    # IDは既存のものを保持するか、新規に生成する
                    id_value = row[0].value if row[0].value else str(uuid.uuid4())
                    name = row[1].value
                    address = row[2].value
                    pickup_time_morning = row[3].value or ""
                    dropoff_time_morning = row[4].value or ""
                    pickup_time_evening = row[5].value or ""
                    dropoff_time_evening = row[6].value or ""
                    constraints = row[7].value or ""
                    attendance_days = [day.strip() for day in (row[8].value or "").split(",")] if row[8].value else []
                    
                    # 必須項目のチェック
                    if not name or not address:
                        continue
                    
                    # 新しい利用者オブジェクトを作成
                    user = User(
                        id=id_value,
                        name=name,
                        address=address,
                        pickup_time_morning=pickup_time_morning,
                        dropoff_time_morning=dropoff_time_morning,
                        pickup_time_evening=pickup_time_evening,
                        dropoff_time_evening=dropoff_time_evening,
                        constraints=constraints,
                        attendance_days=attendance_days
                    )
                    new_data.append(user)
                
                # データを更新
                self.app.user_list = new_data
                
            elif data_type == "vehicle":
                # 新しいデータリストを作成
                new_data = []
                
                for row in data_rows:
                    # IDは既存のものを保持するか、新規に生成する
                    id_value = row[0].value if row[0].value else str(uuid.uuid4())
                    name = row[1].value
                    
                    # 乗車可能人数
                    try:
                        capacity = int(row[2].value) if row[2].value else 0
                    except (ValueError, TypeError):
                        capacity = 0
                    
                    # 必須項目のチェック
                    if not name or capacity <= 0:
                        continue
                    
                    # 新しい車両オブジェクトを作成
                    vehicle = Vehicle(
                        id=id_value,
                        name=name,
                        capacity=capacity
                    )
                    new_data.append(vehicle)
                
                # データを更新
                self.app.vehicle_list = new_data
            
            # 対応するフレームのリストを更新
            if data_type == "staff" and hasattr(self.app, 'staff_frame'):
                self.app.staff_frame.load_staff_list()
            elif data_type == "user" and hasattr(self.app, 'user_frame'):
                self.app.user_frame.load_user_list()
            elif data_type == "vehicle" and hasattr(self.app, 'vehicle_frame'):
                self.app.vehicle_frame.load_vehicle_list()
            
            # データを保存
            self.app.save_all_data()
            messagebox.showinfo("インポート完了", f"{data_type}データのインポートが完了しました。")
            return True
            
        except Exception as e:
            # エラーが発生した場合は元のデータを復元
            if data_type == "staff":
                self.app.staff_list = original_data
            elif data_type == "user":
                self.app.user_list = original_data
            elif data_type == "vehicle":
                self.app.vehicle_list = original_data
                
            messagebox.showerror("エラー", f"データのインポート中にエラーが発生しました: {str(e)}")
            return False 